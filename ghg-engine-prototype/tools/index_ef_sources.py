"""Build a searchable SQLite index for the repository-local EF source library.

This tool indexes source-library evidence under ``../efs``. It deliberately
does not import factors into the active Climate Metrix factor warehouse.

Default run from ``ghg-engine-prototype``::

    uv run python tools/index_ef_sources.py

The generated SQLite file is intended for source discovery, review, and future
mapping work. Active calculation factors still need to pass the existing
``CanonicalFactorRecord`` / ``SQLiteFactorStore`` contracts before publication.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import sqlite3
import sys
import tempfile
import zipfile
from collections.abc import Iterable, Sequence
from dataclasses import dataclass, field
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import openpyxl

SCHEMA_VERSION = 1
DEFAULT_INDEX_RELATIVE_PATH = Path("index") / "ef_library.sqlite"
MAX_XLSX_HEADER_SCAN_ROWS = 25
MAX_COLUMNS_CAPTURED = 200
MAX_TEXT_PREVIEW_LINES = 80
MAX_ZIP_ENTRIES = 500


@dataclass(frozen=True)
class SourceRelease:
    source_id: str
    name: str
    provider: str | None
    release_label: str | None
    source_category: str | None
    candidate_factor_kind: str | None
    review_status: str | None
    license_status: str | None
    home_url: str | None
    local_path: str | None
    file_globs: tuple[str, ...]
    intended_uses: tuple[str, ...]
    active_ingest_status: str | None
    notes: str | None
    manifest_json: str


@dataclass(frozen=True)
class FileRecord:
    file_id: str
    source_id: str
    relative_path: str
    absolute_path: Path
    extension: str
    byte_size: int
    modified_at: str
    sha256: str | None


@dataclass(frozen=True)
class TableRecord:
    table_name: str
    table_kind: str
    row_count: int | None
    column_count: int | None
    header_values: tuple[str, ...] = ()
    sample_rows: tuple[tuple[str, ...], ...] = ()
    extra: dict[str, Any] = field(default_factory=dict)


@dataclass(frozen=True)
class BuildSummary:
    output_path: Path
    source_count: int
    file_count: int
    table_count: int
    unmatched_file_count: int
    fts_enabled: bool

    def as_dict(self) -> dict[str, Any]:
        return {
            "output_path": str(self.output_path),
            "source_count": self.source_count,
            "file_count": self.file_count,
            "table_count": self.table_count,
            "unmatched_file_count": self.unmatched_file_count,
            "fts_enabled": self.fts_enabled,
        }


def default_repo_root() -> Path:
    return Path(__file__).resolve().parents[2]


def default_efs_root() -> Path:
    return default_repo_root() / "efs"


def load_manifest(path: Path) -> tuple[dict[str, Any], list[SourceRelease]]:
    with path.open("r", encoding="utf-8") as handle:
        payload = json.load(handle)
    raw_sources = payload.get("sources")
    if not isinstance(raw_sources, list):
        raise ValueError("manifest must contain a top-level sources list")
    sources = [_source_from_manifest(raw) for raw in raw_sources]
    source_ids = [source.source_id for source in sources]
    duplicate_ids = sorted({source_id for source_id in source_ids if source_ids.count(source_id) > 1})
    if duplicate_ids:
        raise ValueError(f"manifest contains duplicate source_id values: {duplicate_ids}")
    return payload, sources


def _source_from_manifest(raw: Any) -> SourceRelease:
    if not isinstance(raw, dict):
        raise ValueError("each manifest source must be an object")
    source_id = str(raw.get("source_id") or "").strip()
    name = str(raw.get("name") or "").strip()
    file_globs = raw.get("file_globs")
    if not source_id:
        raise ValueError("manifest source missing source_id")
    if not name:
        raise ValueError(f"manifest source {source_id!r} missing name")
    if not isinstance(file_globs, list) or not file_globs:
        raise ValueError(f"manifest source {source_id!r} must define non-empty file_globs")
    intended_uses = raw.get("intended_uses") or []
    return SourceRelease(
        source_id=source_id,
        name=name,
        provider=_optional_str(raw.get("provider")),
        release_label=_optional_str(raw.get("release_label")),
        source_category=_optional_str(raw.get("source_category")),
        candidate_factor_kind=_optional_str(raw.get("candidate_factor_kind")),
        review_status=_optional_str(raw.get("review_status")),
        license_status=_optional_str(raw.get("license_status")),
        home_url=_optional_str(raw.get("home_url")),
        local_path=_optional_str(raw.get("local_path")),
        file_globs=tuple(str(item).replace("\\", "/") for item in file_globs),
        intended_uses=tuple(str(item) for item in intended_uses),
        active_ingest_status=_optional_str(raw.get("active_ingest_status")),
        notes=_optional_str(raw.get("notes")),
        manifest_json=json.dumps(raw, sort_keys=True),
    )


def build_index(
    *,
    efs_root: Path,
    manifest_path: Path,
    output_path: Path,
    hash_files: bool = True,
    include_unmatched: bool = True,
    max_sample_rows: int = 5,
) -> BuildSummary:
    efs_root = efs_root.resolve()
    manifest_path = manifest_path.resolve()
    output_path = output_path.resolve()
    if not efs_root.is_dir():
        raise FileNotFoundError(f"EF source root does not exist: {efs_root}")
    manifest_payload, sources = load_manifest(manifest_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    matched_paths: set[Path] = set()
    source_files: list[tuple[SourceRelease, FileRecord]] = []
    for source in sources:
        paths = _resolve_source_paths(efs_root, source.file_globs)
        matched_paths.update(paths)
        for path in paths:
            source_files.append((source, _file_record(efs_root, source.source_id, path, hash_files=hash_files)))

    unmatched_files: list[FileRecord] = []
    if include_unmatched:
        for path in sorted(p for p in efs_root.rglob("*") if p.is_file()):
            resolved = path.resolve()
            if resolved == output_path:
                continue
            if _is_library_metadata_path(efs_root, resolved):
                continue
            if _is_generated_index_path(efs_root, resolved):
                continue
            if resolved in matched_paths:
                continue
            if resolved == manifest_path:
                continue
            unmatched_files.append(_file_record(efs_root, "_unmatched", resolved, hash_files=hash_files))

    fd, tmp_name = tempfile.mkstemp(prefix="ef_library_", suffix=".sqlite", dir=output_path.parent)
    os.close(fd)
    tmp_path = Path(tmp_name)
    try:
        conn = sqlite3.connect(tmp_path)
        try:
            conn.row_factory = sqlite3.Row
            fts_enabled = _ensure_schema(conn)
            scan_run_id = _start_scan_run(conn, efs_root, manifest_path, manifest_payload)
            source_count = _insert_sources(conn, sources)
            file_count, table_count = _insert_files_and_tables(
                conn,
                source_files,
                scan_run_id=scan_run_id,
                max_sample_rows=max_sample_rows,
                fts_enabled=fts_enabled,
            )
            if include_unmatched and unmatched_files:
                _insert_unmatched_source(conn)
                unmatched_inserted, unmatched_tables = _insert_files_and_tables(
                    conn,
                    [(UNMATCHED_SOURCE, file_record) for file_record in unmatched_files],
                    scan_run_id=scan_run_id,
                    max_sample_rows=max_sample_rows,
                    fts_enabled=fts_enabled,
                )
                file_count += unmatched_inserted
                table_count += unmatched_tables
                source_count += 1
            _finish_scan_run(conn, scan_run_id, file_count=file_count, table_count=table_count)
            conn.commit()
        finally:
            conn.close()
        os.replace(tmp_path, output_path)
    finally:
        if tmp_path.exists():
            tmp_path.unlink()

    return BuildSummary(
        output_path=output_path,
        source_count=source_count,
        file_count=file_count,
        table_count=table_count,
        unmatched_file_count=len(unmatched_files),
        fts_enabled=fts_enabled,
    )


UNMATCHED_SOURCE = SourceRelease(
    source_id="_unmatched",
    name="Unmatched EF source files",
    provider=None,
    release_label=None,
    source_category="unmatched",
    candidate_factor_kind=None,
    review_status="needs_manifest_entry",
    license_status=None,
    home_url=None,
    local_path=None,
    file_globs=(),
    intended_uses=(),
    active_ingest_status="not_mapped",
    notes="Files present under efs but not matched by manifest file_globs.",
    manifest_json="{}",
)


def _ensure_schema(conn: sqlite3.Connection) -> bool:
    conn.executescript(
        """
        PRAGMA foreign_keys = ON;

        CREATE TABLE scan_runs (
            scan_run_id TEXT PRIMARY KEY,
            schema_version INTEGER NOT NULL,
            efs_root TEXT NOT NULL,
            manifest_path TEXT NOT NULL,
            manifest_json TEXT NOT NULL,
            started_at TEXT NOT NULL,
            finished_at TEXT,
            file_count INTEGER,
            table_count INTEGER
        );

        CREATE TABLE source_releases (
            source_id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            provider TEXT,
            release_label TEXT,
            source_category TEXT,
            candidate_factor_kind TEXT,
            review_status TEXT,
            license_status TEXT,
            home_url TEXT,
            local_path TEXT,
            file_globs_json TEXT NOT NULL,
            intended_uses_json TEXT NOT NULL,
            active_ingest_status TEXT,
            notes TEXT,
            manifest_json TEXT NOT NULL
        );

        CREATE TABLE source_files (
            file_id TEXT PRIMARY KEY,
            source_id TEXT NOT NULL,
            scan_run_id TEXT NOT NULL,
            relative_path TEXT NOT NULL,
            extension TEXT NOT NULL,
            byte_size INTEGER NOT NULL,
            modified_at TEXT NOT NULL,
            sha256 TEXT,
            scan_status TEXT NOT NULL,
            scan_error TEXT,
            table_count INTEGER NOT NULL DEFAULT 0,
            FOREIGN KEY(source_id) REFERENCES source_releases(source_id),
            FOREIGN KEY(scan_run_id) REFERENCES scan_runs(scan_run_id)
        );

        CREATE TABLE source_tables (
            table_id TEXT PRIMARY KEY,
            file_id TEXT NOT NULL,
            source_id TEXT NOT NULL,
            table_name TEXT NOT NULL,
            table_kind TEXT NOT NULL,
            row_count INTEGER,
            column_count INTEGER,
            header_json TEXT NOT NULL,
            sample_json TEXT NOT NULL,
            extra_json TEXT NOT NULL,
            FOREIGN KEY(file_id) REFERENCES source_files(file_id),
            FOREIGN KEY(source_id) REFERENCES source_releases(source_id)
        );

        CREATE TABLE source_columns (
            column_id TEXT PRIMARY KEY,
            table_id TEXT NOT NULL,
            file_id TEXT NOT NULL,
            source_id TEXT NOT NULL,
            ordinal INTEGER NOT NULL,
            column_name TEXT NOT NULL,
            normalized_name TEXT NOT NULL,
            inferred_semantic TEXT,
            FOREIGN KEY(table_id) REFERENCES source_tables(table_id),
            FOREIGN KEY(file_id) REFERENCES source_files(file_id),
            FOREIGN KEY(source_id) REFERENCES source_releases(source_id)
        );

        CREATE INDEX idx_source_files_source ON source_files(source_id, relative_path);
        CREATE INDEX idx_source_tables_source ON source_tables(source_id, table_name);
        CREATE INDEX idx_source_columns_semantic ON source_columns(inferred_semantic);
        """
    )
    try:
        conn.execute(
            """
            CREATE VIRTUAL TABLE search_index USING fts5(
                entity_type,
                entity_id,
                source_id,
                content
            )
            """
        )
    except sqlite3.OperationalError:
        return False
    return True


def _start_scan_run(
    conn: sqlite3.Connection,
    efs_root: Path,
    manifest_path: Path,
    manifest_payload: dict[str, Any],
) -> str:
    now = _utc_now()
    scan_run_id = f"scan_{hashlib.sha1(now.encode('utf-8')).hexdigest()[:16]}"
    conn.execute(
        """
        INSERT INTO scan_runs (
            scan_run_id, schema_version, efs_root, manifest_path, manifest_json, started_at
        )
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (
            scan_run_id,
            SCHEMA_VERSION,
            str(efs_root),
            str(manifest_path),
            json.dumps(manifest_payload, sort_keys=True),
            now,
        ),
    )
    return scan_run_id


def _finish_scan_run(
    conn: sqlite3.Connection,
    scan_run_id: str,
    *,
    file_count: int,
    table_count: int,
) -> None:
    conn.execute(
        """
        UPDATE scan_runs
        SET finished_at = ?, file_count = ?, table_count = ?
        WHERE scan_run_id = ?
        """,
        (_utc_now(), file_count, table_count, scan_run_id),
    )


def _insert_sources(conn: sqlite3.Connection, sources: Sequence[SourceRelease]) -> int:
    for source in sources:
        _insert_source(conn, source)
    return len(sources)


def _insert_unmatched_source(conn: sqlite3.Connection) -> None:
    _insert_source(conn, UNMATCHED_SOURCE)


def _insert_source(conn: sqlite3.Connection, source: SourceRelease) -> None:
    conn.execute(
        """
        INSERT OR REPLACE INTO source_releases (
            source_id, name, provider, release_label, source_category, candidate_factor_kind,
            review_status, license_status, home_url, local_path, file_globs_json,
            intended_uses_json, active_ingest_status, notes, manifest_json
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            source.source_id,
            source.name,
            source.provider,
            source.release_label,
            source.source_category,
            source.candidate_factor_kind,
            source.review_status,
            source.license_status,
            source.home_url,
            source.local_path,
            json.dumps(list(source.file_globs), sort_keys=True),
            json.dumps(list(source.intended_uses), sort_keys=True),
            source.active_ingest_status,
            source.notes,
            source.manifest_json,
        ),
    )


def _insert_files_and_tables(
    conn: sqlite3.Connection,
    source_files: Sequence[tuple[SourceRelease, FileRecord]],
    *,
    scan_run_id: str,
    max_sample_rows: int,
    fts_enabled: bool,
) -> tuple[int, int]:
    file_count = 0
    table_count = 0
    for source, file_record in source_files:
        scan_status = "ok"
        scan_error = None
        tables: list[TableRecord] = []
        try:
            tables = list(_scan_file_tables(file_record.absolute_path, max_sample_rows=max_sample_rows))
        except Exception as exc:  # noqa: BLE001 - scanner should preserve errors and keep going
            scan_status = "error"
            scan_error = f"{type(exc).__name__}: {exc}"

        conn.execute(
            """
            INSERT INTO source_files (
                file_id, source_id, scan_run_id, relative_path, extension, byte_size,
                modified_at, sha256, scan_status, scan_error, table_count
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                file_record.file_id,
                source.source_id,
                scan_run_id,
                file_record.relative_path,
                file_record.extension,
                file_record.byte_size,
                file_record.modified_at,
                file_record.sha256,
                scan_status,
                scan_error,
                len(tables),
            ),
        )
        if fts_enabled:
            _insert_search_document(
                conn,
                entity_type="file",
                entity_id=file_record.file_id,
                source_id=source.source_id,
                content=" ".join(
                    part
                    for part in [
                        source.name,
                        source.provider or "",
                        source.release_label or "",
                        file_record.relative_path,
                        source.notes or "",
                    ]
                    if part
                ),
            )
        for table_index, table in enumerate(tables):
            table_id = _stable_id("tbl", file_record.file_id, str(table_index), table.table_name)
            _insert_table(
                conn,
                table,
                table_id=table_id,
                file_id=file_record.file_id,
                source_id=source.source_id,
                fts_enabled=fts_enabled,
            )
            table_count += 1
        file_count += 1
    return file_count, table_count


def _insert_table(
    conn: sqlite3.Connection,
    table: TableRecord,
    *,
    table_id: str,
    file_id: str,
    source_id: str,
    fts_enabled: bool,
) -> None:
    header_json = json.dumps(list(table.header_values), sort_keys=True)
    sample_json = json.dumps([list(row) for row in table.sample_rows], sort_keys=True)
    extra_json = json.dumps(table.extra, sort_keys=True)
    conn.execute(
        """
        INSERT INTO source_tables (
            table_id, file_id, source_id, table_name, table_kind, row_count, column_count,
            header_json, sample_json, extra_json
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            table_id,
            file_id,
            source_id,
            table.table_name,
            table.table_kind,
            table.row_count,
            table.column_count,
            header_json,
            sample_json,
            extra_json,
        ),
    )
    for ordinal, column_name in enumerate(table.header_values):
        normalized = _normalize_column_name(column_name)
        conn.execute(
            """
            INSERT INTO source_columns (
                column_id, table_id, file_id, source_id, ordinal, column_name,
                normalized_name, inferred_semantic
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                _stable_id("col", table_id, str(ordinal), normalized),
                table_id,
                file_id,
                source_id,
                ordinal,
                column_name,
                normalized,
                _infer_semantic(normalized),
            ),
        )
    if fts_enabled:
        content = " ".join(
            [
                table.table_name,
                table.table_kind,
                " ".join(table.header_values),
                " ".join(" ".join(row) for row in table.sample_rows),
                json.dumps(table.extra, sort_keys=True),
            ]
        )
        _insert_search_document(
            conn,
            entity_type="table",
            entity_id=table_id,
            source_id=source_id,
            content=content,
        )


def _insert_search_document(
    conn: sqlite3.Connection,
    *,
    entity_type: str,
    entity_id: str,
    source_id: str,
    content: str,
) -> None:
    conn.execute(
        """
        INSERT INTO search_index (entity_type, entity_id, source_id, content)
        VALUES (?, ?, ?, ?)
        """,
        (entity_type, entity_id, source_id, content),
    )


def _scan_file_tables(path: Path, *, max_sample_rows: int) -> Iterable[TableRecord]:
    extension = path.suffix.lower()
    if extension == ".xlsx":
        yield from _scan_xlsx(path, max_sample_rows=max_sample_rows)
    elif extension == ".csv":
        yield _scan_csv(path, max_sample_rows=max_sample_rows)
    elif extension == ".zip":
        yield from _scan_zip(path)
    elif extension in {".txt", ".md"}:
        yield _scan_text(path)


def _scan_xlsx(path: Path, *, max_sample_rows: int) -> Iterable[TableRecord]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for worksheet in workbook.worksheets:
        header, sample_rows = _worksheet_preview(worksheet, max_sample_rows=max_sample_rows)
        yield TableRecord(
            table_name=worksheet.title,
            table_kind="xlsx_sheet",
            row_count=worksheet.max_row,
            column_count=worksheet.max_column,
            header_values=tuple(header),
            sample_rows=tuple(tuple(row) for row in sample_rows),
            extra={"sheet_state": worksheet.sheet_state},
        )
    workbook.close()


def _worksheet_preview(worksheet: Any, *, max_sample_rows: int) -> tuple[list[str], list[list[str]]]:
    max_row = min(int(worksheet.max_row or 0), MAX_XLSX_HEADER_SCAN_ROWS)
    max_col = min(int(worksheet.max_column or 0), MAX_COLUMNS_CAPTURED)
    if max_row <= 0 or max_col <= 0:
        return [], []

    rows = list(worksheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True))
    header_index = None
    for index, row in enumerate(rows):
        if any(_cell_to_text(cell) for cell in row):
            header_index = index
            break
    if header_index is None:
        return [], []
    header = [_cell_to_text(cell) for cell in rows[header_index]]
    samples: list[list[str]] = []
    for row in rows[header_index + 1 :]:
        rendered = [_cell_to_text(cell) for cell in row]
        if any(rendered):
            samples.append(rendered)
        if len(samples) >= max_sample_rows:
            break
    return header, samples


def _scan_csv(path: Path, *, max_sample_rows: int) -> TableRecord:
    row_count = 0
    header: list[str] = []
    sample_rows: list[list[str]] = []
    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
        reader = csv.reader(handle)
        for row in reader:
            rendered = [_cell_to_text(cell) for cell in row]
            if row_count == 0:
                header = rendered[:MAX_COLUMNS_CAPTURED]
            elif len(sample_rows) < max_sample_rows:
                sample_rows.append(rendered[:MAX_COLUMNS_CAPTURED])
            row_count += 1
    return TableRecord(
        table_name=path.name,
        table_kind="csv",
        row_count=row_count,
        column_count=len(header),
        header_values=tuple(header),
        sample_rows=tuple(tuple(row) for row in sample_rows),
    )


def _scan_zip(path: Path) -> Iterable[TableRecord]:
    with zipfile.ZipFile(path) as archive:
        for info in archive.infolist()[:MAX_ZIP_ENTRIES]:
            yield TableRecord(
                table_name=info.filename,
                table_kind="zip_entry",
                row_count=None,
                column_count=None,
                header_values=(),
                sample_rows=(),
                extra={
                    "compressed_size": info.compress_size,
                    "file_size": info.file_size,
                    "is_dir": info.is_dir(),
                    "date_time": list(info.date_time),
                },
            )


def _scan_text(path: Path) -> TableRecord:
    lines: list[str] = []
    line_count = 0
    with path.open("r", encoding="utf-8", errors="replace") as handle:
        for line in handle:
            if len(lines) < MAX_TEXT_PREVIEW_LINES:
                stripped = line.strip()
                if stripped:
                    lines.append(stripped)
            line_count += 1
    header = lines[:1]
    sample = [[line] for line in lines[1:6]]
    return TableRecord(
        table_name=path.name,
        table_kind="text_preview",
        row_count=line_count,
        column_count=1 if lines else 0,
        header_values=tuple(header),
        sample_rows=tuple(tuple(row) for row in sample),
    )


def _resolve_source_paths(efs_root: Path, file_globs: Sequence[str]) -> list[Path]:
    paths: set[Path] = set()
    for pattern in file_globs:
        paths.update(path.resolve() for path in efs_root.glob(pattern) if path.is_file())
    return sorted(paths)


def _file_record(efs_root: Path, source_id: str, path: Path, *, hash_files: bool) -> FileRecord:
    stat = path.stat()
    relative_path = path.relative_to(efs_root).as_posix()
    return FileRecord(
        file_id=_stable_id("file", relative_path),
        source_id=source_id,
        relative_path=relative_path,
        absolute_path=path,
        extension=path.suffix.lower().lstrip(".") or "no_extension",
        byte_size=stat.st_size,
        modified_at=datetime.fromtimestamp(stat.st_mtime, UTC).isoformat(),
        sha256=_sha256(path) if hash_files else None,
    )


def _is_generated_index_path(efs_root: Path, path: Path) -> bool:
    try:
        rel = path.relative_to(efs_root)
    except ValueError:
        return False
    return bool(rel.parts and rel.parts[0] == "index")


def _is_library_metadata_path(efs_root: Path, path: Path) -> bool:
    try:
        rel = path.relative_to(efs_root)
    except ValueError:
        return False
    return rel.as_posix() in {"README.md", "manifest.json"}


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _stable_id(prefix: str, *parts: str) -> str:
    payload = "|".join(parts)
    return f"{prefix}_{hashlib.sha1(payload.encode('utf-8')).hexdigest()[:16]}"


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return " ".join(text.split())


def _optional_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalize_column_name(value: str) -> str:
    chars = []
    previous_was_sep = False
    for char in value.strip().lower():
        if char.isalnum():
            chars.append(char)
            previous_was_sep = False
        elif not previous_was_sep:
            chars.append("_")
            previous_was_sep = True
    return "".join(chars).strip("_")


def _infer_semantic(normalized: str) -> str | None:
    if not normalized:
        return None
    token = f"_{normalized}_"
    checks = [
        ("factor_value", ("factor", "emission_factor", "ef", "co2e_per")),
        ("unit", ("unit", "uom")),
        ("gas", ("gas", "ghg", "co2", "ch4", "n2o")),
        ("year", ("year", "data_year", "reporting_year")),
        ("geography", ("country", "state", "region", "subregion", "area", "iso")),
        ("category", ("category", "sector", "naics", "bea", "code", "source")),
        ("activity", ("activity", "fuel", "vehicle", "material", "commodity")),
        ("gwp", ("gwp", "ar5", "ar6")),
    ]
    for semantic, needles in checks:
        if any(f"_{needle}_" in token or token.strip("_") == needle for needle in needles):
            return semantic
    return None


def _utc_now() -> str:
    return datetime.now(UTC).isoformat()


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    efs_root = default_efs_root()
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--efs-root",
        type=Path,
        default=efs_root,
        help="Path to the EF source root. Defaults to ../efs from ghg-engine-prototype.",
    )
    parser.add_argument(
        "--manifest",
        type=Path,
        default=None,
        help="Path to manifest JSON. Defaults to <efs-root>/manifest.json.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="SQLite output path. Defaults to <efs-root>/index/ef_library.sqlite.",
    )
    parser.add_argument(
        "--skip-hashes",
        action="store_true",
        help="Skip SHA-256 hashing for faster indexing of large archives.",
    )
    parser.add_argument(
        "--no-unmatched",
        action="store_true",
        help="Do not include files that are present under efs but absent from the manifest globs.",
    )
    parser.add_argument(
        "--max-sample-rows",
        type=int,
        default=5,
        help="Maximum sample rows to keep per CSV/XLSX table.",
    )
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    efs_root = args.efs_root.resolve()
    manifest = (args.manifest or efs_root / "manifest.json").resolve()
    output = (args.output or efs_root / DEFAULT_INDEX_RELATIVE_PATH).resolve()
    summary = build_index(
        efs_root=efs_root,
        manifest_path=manifest,
        output_path=output,
        hash_files=not args.skip_hashes,
        include_unmatched=not args.no_unmatched,
        max_sample_rows=max(0, int(args.max_sample_rows)),
    )
    print(json.dumps(summary.as_dict(), indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    sys.exit(main())
